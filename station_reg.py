#!/usr/bin/python3
import os

import mysql.connector
from mysql.connector import Error
import openpyxl
from dotenv import load_dotenv
load_dotenv()


def create_mysql_connection(host_name, user_name, user_password):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password
        )
    except Error as err:
        print(f"Error: '{err}'")
    return connection


def execute_query(connection, data):
    cursor = connection.cursor()
    query = """
    SELECT ReceivedAt,Message
    FROM CMProdSys.SystemEvents
    WHERE Message LIKE '%{0}%' AND ReceivedAt > '2022-11-08 13:57:00'
    """.format(data)
    try:
        cursor.execute(query)
        row = [[]]
        for (ReceivedAt, Message) in cursor.fetchall():
            rcvd_date = "{:%Y-%m-%d}".format(ReceivedAt)
            rcvd_time = "{:%I:%M%p}".format(ReceivedAt)
            rcvd_message = Message
#            row = rcvd_date, rcvd_time, rcvd_message
            row.append([rcvd_date, rcvd_time, rcvd_message])
        return row
    except Error as err:
        print(f"Error: '{err}'")
        exit(2)


def parse_ext(a_list):
    pass


# MariaDB Host/IP, user and password
db_ip = os.getenv('db_ip')
db_user = os.getenv('db_user')
db_password = os.getenv('db_password')
data = 'ip-a s'
wb = openpyxl.Workbook()
sheet = wb.active

# Initiate database connection
database = create_mysql_connection(db_ip, db_user, db_password)

output = execute_query(database, data)
database.close()

iter_1 = iter_2 = iter_3 = 0

sheet.cell(row=1, column=1).value = 'Date'
sheet.cell(row=1, column=2).value = 'Time'
sheet.cell(row=1, column=3).value = 'Message'

for i in output:
    iter_1 += 1
    iter_2 = 0
    for x in i:
        iter_2 += 1
        if iter_2 == 3:
            x = x.split()
            x = int(x[4])
        sheet.cell(row=iter_1, column=iter_2).value = x

wb.save('station_reg.xlsx')

