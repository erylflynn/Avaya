#!/usr/bin/python3
import paramiko
import time
import re
import os.path
import openpyxl
from dotenv import load_dotenv
load_dotenv()


def cm_connect(host, user, passwd):
    remote_conn_pre = paramiko.SSHClient()
    remote_conn_pre.set_missing_host_key_policy(
        paramiko.AutoAddPolicy())
    remote_conn_pre.connect(host, username=user, port=5022, password=passwd, look_for_keys=False, allow_agent=False)
    remote_conn = remote_conn_pre.invoke_shell()
    time.sleep(1)
    # Terminal type reqeuest OSSI for machine interfaces and OSSIZ for super user
    remote_conn.send("ossiz\n")
    time.sleep(2)
    output = remote_conn.recv(5000)
    ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
    trash = ansi_escape.sub('', output.decode('utf-8'))
    return remote_conn


def list_usage(extension):
    field1 = 'f3e81ff00\n'
    field2 = 'f3e84ff00\n'
    field3 = 'f3e83ff00\n'
    field4 = 'f3e85ff00\n'
    termination = 't\n'
    output_final = ''
    global remote_conn
    command = "clist usage hunt-group " + str(extension) + "\n"
    remote_conn.send(command)
#    time.sleep(.5)
    remote_conn.send(field1)
#    time.sleep(.5)
    remote_conn.send(field2)
#    time.sleep(.5)
    remote_conn.send(field3)
#    time.sleep(.5)
    remote_conn.send(field4)
#    time.sleep(.5)
    remote_conn.send(termination)
    time.sleep(.5)
    while not output_final.endswith('t\n'):
        output = remote_conn.recv(8000)
        ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
        output_final = output_final + ansi_escape.sub('', output.decode('utf-8'))
    output_final = output_final.split('\n')
    output_final.pop(0)
    output_final.pop(0)
    output_final = [item for item in output_final if item != 'n']
    output_final.pop()
    del output_final[-1]
    return output_final


def extension_list(file):
    if not os.path.isfile(file):
        print("Sorry, hunts.txt does not exist.")
        exit()
    usage = []
    with open(file)as f:
        for line in f:
            line.strip()
            line = line.strip('\n')
            usage.append(line)
    return usage


def parse_list_usage(raw_list, extension):
    clean_list = []
    for item in raw_list:
        if item.startswith('d'):
            ext_type = item[1:].split("\t")
            ext_type.insert(0, extension)
            clean_list.append(ext_type)
    return clean_list


def create_list(file):
    extensions = extension_list(file)
    full_list = []
    for ext in extensions:
        full_list.append(parse_list_usage(list_usage(ext), ext))
    return full_list


def list_usage(extension):
    field1 = 'f3e81ff00\n'
    field2 = 'f3e84ff00\n'
    field3 = 'f3e83ff00\n'
    field4 = 'f3e85ff00\n'
    termination = 't\n'
    output_final = ''
    global remote_conn
    command = "clist usage hunt-group " + str(extension) + "\n"
    remote_conn.send(command)
#    time.sleep(.5)
    remote_conn.send(field1)
#    time.sleep(.5)
    remote_conn.send(field2)
#    time.sleep(.5)
    remote_conn.send(field3)
#    time.sleep(.5)
    remote_conn.send(field4)
#    time.sleep(.5)
    remote_conn.send(termination)
    time.sleep(.5)
    while not output_final.endswith('t\n'):
        output = remote_conn.recv(8000)
        ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
        output_final = output_final + ansi_escape.sub('', output.decode('utf-8'))
    output_final = output_final.split('\n')
    output_final.pop(0)
    output_final.pop(0)
    output_final = [item for item in output_final if item != 'n']
    output_final.pop()
    del output_final[-1]
    return output_final


# Add your user name and password for CM
pbx_ip = os.getenv('pbx_ip')
pbx_user = os.getenv('pbx_user')
pbx_password = os.getenv('pbx_password')
list_of_ext = (os.path.dirname(__file__) + '/hunts.txt')
wb = openpyxl.Workbook()
sheet = wb.active

remote_conn = cm_connect(pbx_ip, pbx_user, pbx_password)

list_usage_results = create_list(list_of_ext)

iter_1 = iter_2 = iter_3 = 0

sheet.cell(row=1, column=1).value = 'Skill'
sheet.cell(row=1, column=2).value = 'Used By'
sheet.cell(row=1, column=3).value = 'Ext/Hunt/Vector'
sheet.cell(row=1, column=4).value = 'Aux Value 1'
sheet.cell(row=1, column=5).value = 'Aux Value 2'

iter_1 += 1
for i in list_usage_results:
    for x in i:
        iter_1 += 1
        iter_2 = 0
        for t in x:
            iter_2 += 1
            if t.isdigit():
                t = int(t)
            sheet.cell(row=iter_1, column=iter_2).value = t

wb.save('List_Usage_hunt.xlsx')
