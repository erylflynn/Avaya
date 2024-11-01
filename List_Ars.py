#!/usr/bin/python3
import os

import paramiko
import time
import re
import openpyxl
from dotenv import load_dotenv
load_dotenv()


def cm_connect(host, user, passwd):
    remote_conn_pre = paramiko.SSHClient()
    remote_conn_pre.set_missing_host_key_policy(
        paramiko.AutoAddPolicy())
    remote_conn_pre.connect(host, username=user, port=5022, password=passwd, look_for_keys=False, allow_agent=False)
    remote_conn = remote_conn_pre.invoke_shell()
    time.sleep(1)
    # Terminal type reqeuest OSSI for machine interfaces and OSSIZ for super user
    remote_conn.send("ossiz\n")
    time.sleep(2)
    output = remote_conn.recv(5000)
    ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
    trash = ansi_escape.sub('', output.decode('utf-8'))
    return remote_conn


def list_ars(location):
    field1 = 'f807fff00\n'
    field2 = 'f0003ff00\n'
    termination = 't\n'
    output_final = ''
    global remote_conn
    command = "clist ars analysis location  " + str(location) + "\n"
    remote_conn.send(command)
#    time.sleep(.5)
    remote_conn.send(field1)
#    time.sleep(.5)
    remote_conn.send(field2)
#    time.sleep(.5)
    remote_conn.send(termination)
    time.sleep(.5)
    while not output_final.endswith('t\n'):
        output = remote_conn.recv(8000)
        ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
        output_final = output_final + ansi_escape.sub('', output.decode('utf-8'))
    output_final = output_final.split('\n')
    output_final.pop(0)
    output_final.pop(0)
    output_final = [item for item in output_final if item != 'n']
    output_final.pop()
    del output_final[-1]
    output_final = clean_list(output_final)
    return output_final


def clean_list(raw_ars):
    ars_list = []
    for item in raw_ars:
        if item[0] == "d":
            item = item[1:]
            ars_list.append(item.split())
    return ars_list


def build_list(cleaned_list, location):
    print("Grabbing ARS location " + str(location))
    raw_list = list_ars(location)
    for i in raw_list:
        i.append(location)
        cleaned_list.append(i)
    return cleaned_list


# Add your user name and password for CM
pbx_ip = os.getenv('pbx_ip')
pbx_user = os.getenv('pbx_user')
pbx_password = os.getenv('pbx_password')
wb = openpyxl.Workbook()
sheet = wb.active

remote_conn = cm_connect(pbx_ip, pbx_user, pbx_password)

iter_1 = iter_2 = iter_3 = 1
#list_ars_results = list_ars("all")

full_ars_list = []
full_ars_list = build_list(full_ars_list, "all")
for i in range(1,256):
    full_ars_list = build_list(full_ars_list, i)

sheet.cell(row=1, column=1).value = "Location"
sheet.cell(row=1, column=2).value = "Dial String"
sheet.cell(row=1, column=3).value = "Route Pattern"

for i in full_ars_list:
    iter_1 += 1
    sheet.cell(row=iter_1, column=1).value = i[2]
    sheet.cell(row=iter_1, column=2).value = i[0]
    sheet.cell(row=iter_1, column=3).value = i[1]

wb.save('List_ars.xlsx')
