#!/usr/bin/python3
import paramiko
import time
import re
import os.path
import openpyxl
from dotenv import load_dotenv
load_dotenv()


def cm_connect(host, user, passwd):
    remote_conn_pre = paramiko.SSHClient()
    remote_conn_pre.set_missing_host_key_policy(
        paramiko.AutoAddPolicy())
    remote_conn_pre.connect(host, username=user, port=5022, password=passwd, look_for_keys=False, allow_agent=False)
    remote_conn = remote_conn_pre.invoke_shell()
    time.sleep(1)
    # Terminal type reqeuest OSSI for machine interfaces and OSSIZ for super user
    remote_conn.send("ossiz\n")
    time.sleep(2)
    output = remote_conn.recv(5500)
    ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
    trash = ansi_escape.sub('', output.decode('utf-8'))
    return remote_conn


def list_station_port(card):
    field = 'f8005ff00\n'
    termination = 't\n'
    output_final = ''
    global remote_conn
    command = "clist station port " + card + "\n"
    remote_conn.send(command)
    time.sleep(1)
    remote_conn.send(field)
    time.sleep(1)
    remote_conn.send(termination)
    time.sleep(1)
    while not output_final.endswith('t\n'):
        output = remote_conn.recv(8000)
        ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
        output_final = output_final + ansi_escape.sub('', output.decode('utf-8'))
    log = open((os.path.dirname(__file__) + '\log.txt'), 'w')
    output_final = output_final.split('\n')
    log.write(str(output_final))
    output_final.pop(0)
    output_final.pop(0)
    output_final = [item for item in output_final if item != 'n']
    output_final.pop()
    del output_final[-1]
    return output_final


def status_station(extension):
    field_ext = 'f0004ff00\n'
    field_network_region = 'f1fa4ff00\n'
    termination = 't\n'
    output_final = ''
    global remote_conn
    command = "cstatus station " + extension + "\n"
    remote_conn.send(command)
    time.sleep(1)
    remote_conn.send(field_ext)
    time.sleep(1)
    remote_conn.send(field_network_region)
    time.sleep(1)
    remote_conn.send(termination)
    time.sleep(1)
    while not output_final.endswith('t\n'):
        output = remote_conn.recv(8000)
        ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
        output_final = output_final + ansi_escape.sub('', output.decode('utf-8'))
    #log = open((os.path.dirname(__file__) + '\log.txt'), 'w')
    output_final = output_final.split('\n')
    #print(output_final)
    #log.write(str(output_final))
    maint = 'All maintenance resources busy'
    if maint in output_final[1]:
        return 'bad'
    output_final.pop(0)
    output_final.pop(0)
    output_final = [item for item in output_final if item != 'n']
    output_final.pop()
    del output_final[-1]
    #print(output_final)
    return output_final

def parse_phone_list(cards):
    phone_list = []
    status_list = []
    for card in cards:
        phone_list = list_station_port(card)
        for extension in phone_list:
            extension = extension[1:]
            results = 'bad'
            while results == 'bad':
                results = str(status_station(extension))
            results = str(results[3:])
            results = results[:-2]
            results = results.split('\\t')
            status = results[0]
            network_region = (results[1])
            #print(status + " " + network_region)
            status_list.append([extension,status,network_region])
    return status_list


def parse_line_cards(line_cards):
    if not os.path.isfile(line_cards):
        print("Sorry, dig_media_gateway.txt does not exist.")
        exit()
    card_list = []
    with open(line_cards) as f:
        for line in f:
            line.strip()
            line = line.strip('\n')
            card_list.append(line)
    return card_list


# Add your username and password for CM
pbx_ip = os.getenv('pbx_ip')
pbx_user = os.getenv('pbx_user')
pbx_password = os.getenv('pbx_password')
list_of_mm711 = (os.path.dirname(__file__) + '\dig_media_gateway.txt')
wb = openpyxl.Workbook()
sheet = wb.active

remote_conn = cm_connect(pbx_ip, pbx_user, pbx_password)

list_of_cards = parse_line_cards(list_of_mm711)
active_stations = parse_phone_list(list_of_cards)


iter_1 = iter_2 = iter_3 = iter_4 = 0

sheet.cell(row=1, column=1).value = 'Extension'
sheet.cell(row=1, column=2).value = 'Service State'
sheet.cell(row=1, column=3).value = 'Location'

iter_1 += 1
for i in active_stations:
    iter_1 += 1
    iter_2 = 0
    for X in i:
        iter_2 += 1
        if X.isdigit():
            X = int(X)
        sheet.cell(row=iter_1, column=iter_2).value = X

wb.save('status_station.xlsx')
